# -*- coding: utf-8 -*-
"""
Created on Mon Apr 12 16:49:09 2021

@author: Administrator
"""


#this website is called macrotrends
#this script is designed to scrape its financial statements
#yahoo finance only contains the recent 5 year
#macrotrends can trace back to 2005 if applicable
import re
import json
import pandas as pd
import requests
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers


#simply scrape
def scrape(url,**kwargs):
    
    session=requests.Session()
    session.headers.update(
            {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36'})
    
    response=session.get(url,**kwargs)

    return response


#create dataframe
def etl(response):

    #regex to find the data
    num=re.findall('(?<=div\>\"\,)[0-9\.\"\:\-\, ]*',response.text)
    text=re.findall('(?<=s\: \')\S+(?=\'\, freq)',response.text)

    #convert text to dict via json
    dicts=[json.loads('{'+i+'}') for i in num]

    #create dataframe
    df=pd.DataFrame()
    for ind,val in enumerate(text):
        df[val]=dicts[ind].values()
    df.index=dicts[ind].keys()
    
    return df

# Function to clean up row index
def clean_row_index(df):
    df.index = df.index.str.replace('-', ' ').str.title()
    return df

# Function to convert strings to numeric data type
def convert_to_numeric(df):
    df = df.apply(pd.to_numeric, errors='ignore')
    return df

# Function to align text to the left in the first column of each sheet
def align_text(filename, sheet_names):
    wb = load_workbook(filename)
    for sheet in sheet_names:
        ws = wb[sheet]
        for cell in ws['A']:
            cell.alignment = Alignment(horizontal='left')
    wb.save(filename)

# Function to format numbers
def format_numbers(filename, sheet_names):
    wb = load_workbook(filename)
    for sheet in sheet_names:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER_00
    wb.save(filename)

# Function to autofit column widths
def autofit_column_widths(filename, sheet_names):
    wb = load_workbook(filename)
    for sheet in sheet_names:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    wb.save(filename)


def main(tickers):
    # Create a new directory for the files
    os.makedirs('financial_statements', exist_ok=True)

    for ticker, company in tickers.items():
        urls = [
            # Links to annual financial statements
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/income-statements?freq=A',
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/balance-sheet?freq=A',
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/cash-flow-statement?freq=A',
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/financial-ratios?freq=A',
            # Links to quarterly financial statements
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/income-statement?freq=Q',
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/balance-sheet?freq=Q',
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/cash-flow-statement?freq=Q',
            f'https://www.macrotrends.net/stocks/charts/{ticker}/{company}/financial-ratios?freq=Q'
        ]
        sheet_names = ['Income Statement Annually', 'Balance Sheet Annually', 'Cash Flow Statement Annually', 'Key Metrics Annually',
                       'Income Statement Quarterly', 'Balance Sheet Quarterly', 'Cash Flow Statement Quarterly', 'Key Metrics Quarterly'
                       ]
        filename = f'financial_statements/{ticker}_financial_statements.xlsx'
        
        with pd.ExcelWriter(filename) as writer:
            for url, sheet_name in zip(urls, sheet_names):
                response=scrape(url)
                df=etl(response)
                df = df.transpose()  # Transpose the dataframe
                df = clean_row_index(df)  # Clean up row index
                df = convert_to_numeric(df)  # Convert strings to numeric data type
                df.to_excel(writer, sheet_name=sheet_name)
        
        align_text(filename, sheet_names)
        format_numbers(filename, sheet_names)
        autofit_column_widths(filename, sheet_names)
    
    return


if __name__ == "__main__":
    tickers = {'AAPL': 'apple', 'PLTR': 'palantir-technologies', 'MSFT': 'microsoft', 'GOOG': 'google'}  # Add more tickers as needed
    main(tickers)
